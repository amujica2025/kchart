import pandas as pd
import os
import glob
from openpyxl import load_workbook

# Path to your files
folder_path = r'C:\Users\alexm\Downloads\kbot\my-kalshi-app\csv data\match_price_history'
file_pattern = os.path.join(folder_path, "*.xlsx")
files = glob.glob(file_pattern)

# Define the headers and formulas
headers = ["starting", "ending", "min", "max", "median", "stddev"]
# Note: Excel formulas use column letters. 
# B20 is column B, C20 is column C, etc.
formulas = [
    ["=B20", "=LOOKUP(1.E+308, B:B)", "=MIN(B20:B)", "=MAX(B20:B)", "=MEDIAN(B20:B)", "=STDEV.S(B20:B)"],
    ["=C20", "=LOOKUP(1.E+308, C:C)", "=MIN(C20:C)", "=MAX(C20:C)", "=MEDIAN(C20:C)", "=STDEV.S(C20:C)"],
    ["=D20", "=LOOKUP(1.E+308, D:D)", "=MIN(D20:D)", "=MAX(D20:D)", "=MEDIAN(D20:D)", "=STDEV.S(D20:D)"],
    ["=E20", "=LOOKUP(1.E+308, E:E)", "=MIN(E20:E)", "=MAX(E20:E)", "=MEDIAN(E20:E)", "=STDEV.S(E20:E)"],
    ["=F20", "=LOOKUP(1.E+308, F:F)", "=MIN(F20:F)", "=MAX(F20:F)", "=MEDIAN(F20:F)", "=STDEV.S(F20:F)"]
]

for file in files:
    try:
        wb = load_workbook(file)
        ws = wb.active # Assumes the data is on the first sheet

        # Write Headers (Row 1)
        for col_idx, header in enumerate(headers, start=7): # Column G is 7
            ws.cell(row=1, column=col_idx, value=header)

        # Write Formulas (Rows 2-4)
        for row_idx in range(2, 5): 
            for col_idx, formula in enumerate(formulas, start=7):
                ws.cell(row=row_idx, column=col_idx, value=formula[col_idx - 7])

        wb.save(file)
        print(f"Processed: {os.path.basename(file)}")
    except Exception as e:
        print(f"Error processing {file}: {e}")

print("Batch update complete.")