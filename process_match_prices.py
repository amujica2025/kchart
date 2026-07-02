#!/usr/bin/env python3
from __future__ import annotations

import csv, json, math
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

try:
    import numpy as np
except Exception:
    np = None

INPUT_DIR = Path(r"C:\Users\alexm\Downloads\kbot\my-kalshi-app\csv data\match_price_history")
PREFIX = "kalshi-price-history"

OUT_CSV = INPUT_DIR / "csvprice data"
OUT_NPY = INPUT_DIR / "npy data"
OUT_JSON = INPUT_DIR / "json data"

for p in (OUT_CSV, OUT_NPY, OUT_JSON):
    p.mkdir(parents=True, exist_ok=True)


def safe_float(v):
    if v in (None, ""):
        return None
    try:
        x = float(v)
        if math.isnan(x):
            return None
        return x
    except Exception:
        return None


def parse_time(v):
    if isinstance(v, datetime):
        return v
    try:
        return datetime.fromisoformat(str(v).replace("Z", "+00:00"))
    except Exception:
        return str(v)


def iso_time(v):
    return v.isoformat() if hasattr(v, "isoformat") else str(v)


def price_code(price):
    cents = int(round(float(price)))
    cents = max(0, min(100, cents))
    return f"{cents:04d}"


def suffix_from_index(i):
    letters = "abcdefghijklmnopqrstuvwxyz"
    i += 1
    out = ""
    while i:
        i -= 1
        out = letters[i % 26] + out
        i //= 26
    return out


def trim_edge_filler(data):
    """
    Removes repeated flat filler at beginning/end.
    Keeps one anchor row before first movement and one final row after last movement.
    """
    if len(data) <= 2:
        return data, 0, 0

    first_fav, first_dog = data[0][1], data[0][2]
    start = 0

    for i in range(1, len(data)):
        if data[i][1] != first_fav or data[i][2] != first_dog:
            start = max(0, i - 1)
            break

    last_fav, last_dog = data[-1][1], data[-1][2]
    end = len(data) - 1

    for i in range(len(data) - 2, -1, -1):
        if data[i][1] != last_fav or data[i][2] != last_dog:
            end = min(len(data) - 1, i + 1)
            break

    return data[start:end + 1], start, len(data) - 1 - end


def read_workbook(path: Path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    header_row = None
    headers = None

    for r in range(1, min(50, ws.max_row) + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, 4)]
        if str(row_vals[0]).strip().lower() == "timestamp":
            header_row = r
            headers = row_vals
            break

    if header_row is None:
        raise ValueError("timestamp header not found")

    # Your workbook format:
    # timestamp | favorite player | dog player
    fav_player = str(headers[1]).strip() if headers[1] else "Favorite"
    dog_player = str(headers[2]).strip() if headers[2] else "Dog"

    event_ticker = ""
    final_score = ""
    match_winner = ""
    kalshi_date = ""

    for r in range(1, header_row):
        v = ws.cell(r, 1).value
        if not isinstance(v, str):
            continue

        if v.startswith("# event_ticker="):
            event_ticker = v.split("=", 1)[1].strip()
        elif v.startswith("# final_score="):
            final_score = v.split("=", 1)[1].strip()
        elif v.startswith("# match_winner="):
            match_winner = v.split("=", 1)[1].strip()
        elif v.startswith("# kalshi_date="):
            kalshi_date = v.split("=", 1)[1].strip()

    raw = []

    for row in ws.iter_rows(min_row=header_row + 1, max_col=3, values_only=True):
        ts = row[0]
        fav_price = safe_float(row[1])
        dog_price = safe_float(row[2])

        if ts in (None, ""):
            continue

        raw.append((parse_time(ts), fav_price, dog_price))

    clean = []
    last_fav = None
    last_dog = None

    for ts, fav_price, dog_price in raw:
        if fav_price is not None:
            last_fav = fav_price
        if dog_price is not None:
            last_dog = dog_price

        if last_fav is not None and last_dog is not None:
            clean.append((ts, float(last_fav), float(last_dog)))

    try:
        wb.close()
    except Exception:
        pass

    if len(clean) < 2:
        raise ValueError("not enough usable rows")

    return {
        "source_workbook": path.name,
        "fav_player": fav_player,
        "dog_player": dog_player,
        "match_winner": match_winner,
        "final_score": final_score,
        "event_ticker": event_ticker,
        "kalshi_date": kalshi_date,
        "data": clean,
    }


def write_player_files(base, info, player_name, opponent_name, side, times, player_prices, opponent_prices, trim_start, trim_end):
    csv_path = OUT_CSV / f"{base}.csv"
    json_path = OUT_JSON / f"{base}.json"
    npy_path = OUT_NPY / f"{base}.npy"

    with csv_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["timestamp", "player_price", "opponent_price"])
        for ts, p, o in zip(times, player_prices, opponent_prices):
            w.writerow([iso_time(ts), p, o])

    payload = {
        "id": base,
        "source_workbook": info["source_workbook"],
        "event_ticker": info["event_ticker"],
        "kalshi_date": info["kalshi_date"],
        "player_name": player_name,
        "opponent_name": opponent_name,
        "player_side": side,
        "match_winner": info["match_winner"],
        "final_score": info["final_score"],
        "starting_price": player_prices[0],
        "ending_price": player_prices[-1],
        "min_price": min(player_prices),
        "max_price": max(player_prices),
        "opponent_starting_price": opponent_prices[0],
        "opponent_ending_price": opponent_prices[-1],
        "row_count": len(player_prices),
        "trimmed_opening_rows": trim_start,
        "trimmed_ending_rows": trim_end,
        "start_timestamp": iso_time(times[0]),
        "end_timestamp": iso_time(times[-1]),
        "files": {
            "csv": f"{base}.csv",
            "json": f"{base}.json",
            "npy": f"{base}.npy",
        },
    }

    json_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")

    if np is not None:
        arr = {
            "timestamp": np.array([iso_time(t) for t in times], dtype=object),
            "player_price": np.array(player_prices, dtype=np.float32),
            "opponent_price": np.array(opponent_prices, dtype=np.float32),
        }
        np.save(npy_path, np.array(arr, dtype=object), allow_pickle=True)


def main():
    files = sorted(INPUT_DIR.glob(PREFIX + "*.xlsx"))

    if not files:
        print("No matching workbooks found.")
        return

    suffix_counts = defaultdict(int)
    ok_players = 0
    failed = 0

    print(f"Found {len(files)} workbooks")
    print(f"CSV : {OUT_CSV}")
    print(f"NPY : {OUT_NPY}")
    print(f"JSON: {OUT_JSON}")
    print()

    for idx, path in enumerate(files, 1):
        try:
            info = read_workbook(path)
            trimmed, trim_start, trim_end = trim_edge_filler(info["data"])

            times = [r[0] for r in trimmed]
            fav_prices = [r[1] for r in trimmed]
            dog_prices = [r[2] for r in trimmed]

            players = [
                (info["fav_player"], info["dog_player"], "fav", fav_prices, dog_prices),
                (info["dog_player"], info["fav_player"], "dog", dog_prices, fav_prices),
            ]

            for player_name, opponent_name, side, p_prices, o_prices in players:
                code = price_code(p_prices[0])
                suffix = suffix_from_index(suffix_counts[code])
                suffix_counts[code] += 1
                base = code + suffix

                write_player_files(
                    base,
                    info,
                    player_name,
                    opponent_name,
                    side,
                    times,
                    p_prices,
                    o_prices,
                    trim_start,
                    trim_end,
                )

                ok_players += 1

            print(f"[{idx}/{len(files)}] OK {path.name} -> 2 player files")

        except Exception as e:
            failed += 1
            print(f"[{idx}/{len(files)}] FAIL {path.name}: {e}")

    print()
    print("DONE")
    print(f"Player datasets created: {ok_players}")
    print(f"Failed workbooks: {failed}")

    if np is None:
        print("NOTE: numpy not installed, so .npy files were skipped.")
        print("Install with: pip install numpy")


if __name__ == "__main__":
    main()
